import re
from collections.abc import Iterable, Sequence
from dataclasses import dataclass
from io import BytesIO
from typing import Literal, Protocol, runtime_checkable
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.workbook.workbook import Workbook

from ..domain.errors import PriceListError, PriceListErrorCode
from ..domain.models import PriceListItem
from ..domain.normalization import (
    map_availability,
    normalize_persian_text,
    normalize_product_code,
    normalize_product_name,
    parse_pack_size,
    parse_price_irr,
)

_HEADER_CLEANER = re.compile(r"[\s()\[\]{}_\-/:]+")
_REQUIRED_COLUMNS = {"product_code", "product_name", "price_irr"}


@runtime_checkable
class _ReadOnlySheet(Protocol):
    title: str

    @property
    def max_row(self) -> int | None: ...

    @property
    def max_column(self) -> int | None: ...

    def iter_rows(
        self,
        min_row: int | None = None,
        max_row: int | None = None,
        min_col: int | None = None,
        max_col: int | None = None,
        *,
        values_only: Literal[True],
    ) -> Iterable[tuple[object, ...]]: ...


_HEADER_ALIASES_RAW: dict[str, tuple[str, ...]] = {
    "row_number": ("ردیف", "شماره"),
    "product_code": ("کد کالا", "کد محصول", "شناسه کالا", "sku", "item code"),
    "product_name": ("شرح کالا", "نام کالا", "شرح محصول", "نام محصول", "product"),
    "brand": ("برند", "brand"),
    "unit": ("واحد فروش", "واحد", "unit"),
    "pack_size": (
        "تعداد در بسته",
        "تعداد بسته",
        "تعداد در کارتن",
        "بسته بندی",
        "pack size",
    ),
    "price_irr": (
        "قیمت همکار ریال",
        "قیمت همکار (ریال)",
        "قیمت ریال",
        "قیمت خرید ریال",
        "قیمت",
        "price irr",
    ),
    "availability": ("وضعیت موجودی", "موجودی", "availability", "stock"),
    "notes": ("توضیحات", "توضیح", "یادداشت", "notes"),
}


def _normalize_header(value: object) -> str:
    return _HEADER_CLEANER.sub("", normalize_persian_text(value).casefold())


_HEADER_ALIASES = {
    canonical: {_normalize_header(alias) for alias in aliases}
    for canonical, aliases in _HEADER_ALIASES_RAW.items()
}


@dataclass(frozen=True, slots=True)
class XlsxExtractionLimits:
    max_header_scan_rows: int = 20
    max_rows: int = 50_000
    max_columns: int = 200
    max_sheets: int = 20
    max_archive_entries: int = 10_000
    max_uncompressed_bytes: int = 100 * 1024 * 1024
    max_compression_ratio: float = 250.0

    def __post_init__(self) -> None:
        values = (
            self.max_header_scan_rows,
            self.max_rows,
            self.max_columns,
            self.max_sheets,
            self.max_archive_entries,
            self.max_uncompressed_bytes,
        )
        if any(value < 1 for value in values) or self.max_compression_ratio < 1:
            raise ValueError("XLSX extraction limits must be positive")


class XlsxPriceListExtractor:
    def __init__(self, limits: XlsxExtractionLimits | None = None) -> None:
        self._limits = limits or XlsxExtractionLimits()

    def extract(self, content: bytes, *, filename: str) -> list[PriceListItem]:
        if not content.startswith(b"PK\x03\x04"):
            raise PriceListError(
                PriceListErrorCode.INVALID_XLSX_SIGNATURE,
                "محتوای فایل با فرمت XLSX معتبر مطابقت ندارد.",
                {"filename": filename},
            )

        self._validate_archive(content, filename=filename)

        try:
            workbook = load_workbook(
                filename=BytesIO(content),
                read_only=True,
                data_only=True,
                keep_links=False,
            )
        except (BadZipFile, InvalidFileException, OSError, ValueError) as exc:
            raise PriceListError(
                PriceListErrorCode.WORKBOOK_UNREADABLE,
                "فایل Excel قابل خواندن نیست یا آسیب دیده است.",
                {"filename": filename},
            ) from exc

        try:
            self._validate_workbook_dimensions(workbook)
            sheet, header_row_number, columns, header_values = self._find_header(workbook)
            return self._extract_rows(
                sheet=sheet,
                header_row_number=header_row_number,
                columns=columns,
                header_values=header_values,
            )
        finally:
            workbook.close()

    def _validate_archive(self, content: bytes, *, filename: str) -> None:
        try:
            with ZipFile(BytesIO(content)) as archive:
                entries = archive.infolist()
        except BadZipFile as exc:
            raise PriceListError(
                PriceListErrorCode.WORKBOOK_UNREADABLE,
                "فایل Excel قابل خواندن نیست یا آسیب دیده است.",
                {"filename": filename},
            ) from exc

        if len(entries) > self._limits.max_archive_entries:
            raise self._archive_limit_error(filename)

        total_uncompressed = 0
        for entry in entries:
            if entry.flag_bits & 0x1:
                raise PriceListError(
                    PriceListErrorCode.ENCRYPTED_XLSX_NOT_SUPPORTED,
                    "فایل Excel رمزگذاری‌شده در این نسخه پشتیبانی نمی‌شود.",
                    {"filename": filename},
                )

            total_uncompressed += entry.file_size
            if total_uncompressed > self._limits.max_uncompressed_bytes:
                raise self._archive_limit_error(filename)

            if entry.file_size == 0:
                continue
            if entry.compress_size == 0:
                raise self._archive_limit_error(filename)
            compression_ratio = entry.file_size / entry.compress_size
            if compression_ratio > self._limits.max_compression_ratio:
                raise self._archive_limit_error(filename)

    def _archive_limit_error(self, filename: str) -> PriceListError:
        return PriceListError(
            PriceListErrorCode.XLSX_ARCHIVE_LIMIT_EXCEEDED,
            "ساختار فشرده فایل Excel از محدودیت ایمن پردازش عبور کرده است.",
            {"filename": filename},
        )

    def _validate_workbook_dimensions(self, workbook: Workbook) -> None:
        worksheets = self._read_only_worksheets(workbook)
        if len(worksheets) > self._limits.max_sheets:
            raise PriceListError(
                PriceListErrorCode.WORKBOOK_SHEET_LIMIT_EXCEEDED,
                "تعداد برگه‌های فایل Excel از محدودیت مجاز بیشتر است.",
                {"max_sheets": self._limits.max_sheets},
            )

        max_sheet_rows = self._limits.max_rows + self._limits.max_header_scan_rows
        for sheet in worksheets:
            if (sheet.max_row or 0) > max_sheet_rows:
                raise PriceListError(
                    PriceListErrorCode.WORKBOOK_ROW_LIMIT_EXCEEDED,
                    "تعداد ردیف‌های فایل Excel از محدودیت مجاز بیشتر است.",
                    {"sheet": sheet.title, "max_rows": self._limits.max_rows},
                )
            if (sheet.max_column or 0) > self._limits.max_columns:
                raise PriceListError(
                    PriceListErrorCode.WORKBOOK_COLUMN_LIMIT_EXCEEDED,
                    "تعداد ستون‌های فایل Excel از محدودیت مجاز بیشتر است.",
                    {"sheet": sheet.title, "max_columns": self._limits.max_columns},
                )

    def _find_header(
        self,
        workbook: Workbook,
    ) -> tuple[_ReadOnlySheet, int, dict[str, int], tuple[object, ...]]:
        best: tuple[_ReadOnlySheet, int, dict[str, int], tuple[object, ...]] | None = None

        for sheet in self._read_only_worksheets(workbook):
            scan_limit = min(
                sheet.max_row or self._limits.max_header_scan_rows,
                self._limits.max_header_scan_rows,
            )
            for row_number, row in enumerate(
                sheet.iter_rows(min_row=1, max_row=scan_limit, values_only=True),
                start=1,
            ):
                if any("تومان" in normalize_persian_text(value) for value in row if value):
                    raise PriceListError(
                        PriceListErrorCode.TOMAN_NOT_SUPPORTED,
                        "این فایل برچسب تومان دارد؛ ورودی MVP باید ریال باشد.",
                        {"sheet": sheet.title, "row": row_number},
                    )
                mapping = self._map_columns(row)
                candidate = (sheet, row_number, mapping, tuple(row))
                if best is None or len(mapping) > len(best[2]):
                    best = candidate
                if _REQUIRED_COLUMNS.issubset(mapping):
                    return candidate

        if best is None:
            raise PriceListError(
                PriceListErrorCode.HEADER_NOT_FOUND,
                "سطر عنوان جدول در فایل پیدا نشد.",
            )

        missing = sorted(_REQUIRED_COLUMNS.difference(best[2]))
        raise PriceListError(
            PriceListErrorCode.REQUIRED_COLUMN_MISSING,
            "ستون‌های اجباری کد کالا، شرح کالا و قیمت ریالی پیدا نشدند.",
            {"missing_columns": missing},
        )

    def _map_columns(self, row: Sequence[object]) -> dict[str, int]:
        mapping: dict[str, int] = {}
        for index, value in enumerate(row):
            normalized = _normalize_header(value)
            if not normalized:
                continue
            for canonical, aliases in _HEADER_ALIASES.items():
                if canonical not in mapping and normalized in aliases:
                    mapping[canonical] = index
                    break
        return mapping

    def _extract_rows(
        self,
        *,
        sheet: _ReadOnlySheet,
        header_row_number: int,
        columns: dict[str, int],
        header_values: tuple[object, ...],
    ) -> list[PriceListItem]:
        items: list[PriceListItem] = []
        seen_codes: set[str] = set()

        if sheet.max_row is not None and sheet.max_row - header_row_number > self._limits.max_rows:
            raise PriceListError(
                PriceListErrorCode.WORKBOOK_ROW_LIMIT_EXCEEDED,
                "تعداد ردیف‌های فایل Excel از محدودیت مجاز بیشتر است.",
                {"sheet": sheet.title, "max_rows": self._limits.max_rows},
            )

        rows: Iterable[tuple[object, ...]] = sheet.iter_rows(
            min_row=header_row_number + 1,
            values_only=True,
        )
        for source_row_number, row in enumerate(rows, start=header_row_number + 1):
            if source_row_number - header_row_number > self._limits.max_rows:
                raise PriceListError(
                    PriceListErrorCode.WORKBOOK_ROW_LIMIT_EXCEEDED,
                    "تعداد ردیف‌های فایل Excel از محدودیت مجاز بیشتر است.",
                    {"sheet": sheet.title, "max_rows": self._limits.max_rows},
                )
            if self._row_is_empty(row):
                continue

            raw_code = self._cell(row, columns["product_code"])
            raw_name = self._cell(row, columns["product_name"])
            raw_price = self._cell(row, columns["price_irr"])

            try:
                code_raw_text = self._required_text(
                    raw_code,
                    code=PriceListErrorCode.EMPTY_PRODUCT_CODE,
                )
                product_code = normalize_product_code(raw_code)
                name_raw_text = self._required_text(
                    raw_name,
                    code=PriceListErrorCode.EMPTY_PRODUCT_NAME,
                )
                product_name = normalize_product_name(raw_name)
                price_irr = parse_price_irr(raw_price)
                pack_size = parse_pack_size(self._optional_cell(row, columns, "pack_size"))
            except PriceListError as exc:
                exc.with_details(sheet=sheet.title, row=source_row_number)
                raise

            if product_code in seen_codes:
                raise PriceListError(
                    PriceListErrorCode.DUPLICATE_PRODUCT_CODE,
                    f"کد کالا در فایل تکرار شده است: {product_code}",
                    {
                        "sheet": sheet.title,
                        "row": source_row_number,
                        "product_code": product_code,
                    },
                )
            seen_codes.add(product_code)

            availability_value = self._optional_cell(row, columns, "availability")
            item = PriceListItem(
                source_row_number=source_row_number,
                product_code_raw=code_raw_text,
                product_code_normalized=product_code,
                product_name_raw=name_raw_text,
                product_name_normalized=product_name,
                brand=self._optional_text(self._optional_cell(row, columns, "brand")),
                unit=self._optional_text(self._optional_cell(row, columns, "unit")),
                pack_size=pack_size,
                price_irr=price_irr,
                availability=map_availability(availability_value),
                availability_raw=self._optional_text(availability_value),
                notes=self._optional_text(self._optional_cell(row, columns, "notes")),
                raw_row=self._raw_row(header_values, row),
            )
            items.append(item)

        if not items:
            raise PriceListError(
                PriceListErrorCode.NO_PRODUCT_ROWS,
                "هیچ ردیف کالای معتبری در فایل پیدا نشد.",
            )
        return items

    @staticmethod
    def _cell(row: Sequence[object], index: int) -> object:
        return row[index] if index < len(row) else None

    def _optional_cell(
        self,
        row: Sequence[object],
        columns: dict[str, int],
        name: str,
    ) -> object:
        index = columns.get(name)
        return None if index is None else self._cell(row, index)

    @staticmethod
    def _required_text(value: object, *, code: PriceListErrorCode) -> str:
        text = normalize_persian_text(value)
        if not text:
            message = (
                "کد کالا خالی است."
                if code is PriceListErrorCode.EMPTY_PRODUCT_CODE
                else "شرح کالا خالی است."
            )
            raise PriceListError(code, message)
        return text

    @staticmethod
    def _optional_text(value: object) -> str | None:
        text = normalize_persian_text(value)
        return text or None

    @staticmethod
    def _row_is_empty(row: Sequence[object]) -> bool:
        return not any(value is not None and str(value).strip() for value in row)

    @staticmethod
    def _raw_row(headers: tuple[object, ...], row: Sequence[object]) -> dict[str, object]:
        raw: dict[str, object] = {}
        for index, value in enumerate(row):
            if value is None:
                continue
            header = normalize_persian_text(headers[index]) if index < len(headers) else ""
            raw[header or f"column_{index + 1}"] = value
        return raw

    @staticmethod
    def _read_only_worksheets(workbook: Workbook) -> tuple[_ReadOnlySheet, ...]:
        worksheets = tuple(workbook.worksheets)
        if not workbook.read_only or not all(
            isinstance(sheet, _ReadOnlySheet) for sheet in worksheets
        ):
            raise RuntimeError("Workbook was not opened in read-only mode")
        return tuple(sheet for sheet in worksheets if isinstance(sheet, _ReadOnlySheet))
